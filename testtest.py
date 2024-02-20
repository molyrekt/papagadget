import telebot
import csv
from telebot import types
import openpyxl
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from io import BytesIO

bot = telebot.TeleBot('6709849437:AAHokC-yIJ9S4PwvIWMx8xyK00mcRxVeqEM')





######## ВОЗВРАЩЕНИЕ В ТРЕЙД ИН МЕНЮ ######



def main_menu_tradein(callback):
    file_path = './ppg_tradein.png'
    with open(file_path, 'rb') as image_file:
        markup = types.InlineKeyboardMarkup()

        with open('tradein.csv', 'r') as csv_file:
            csv_reader = csv.reader(csv_file)
            for row in csv_reader:
                _, button_text, callback_data = row
                button = types.InlineKeyboardButton(button_text, callback_data=callback_data)
                markup.add(button)

        bot.send_photo(callback.message.chat.id, image_file, reply_markup=markup)


########## ВОЗВРАЩЕНИЕ В МЕНЮ РЕМОНТНЫХ УСЛУГ ПО 14 ПРО МАКСУ #####


def main_menu_repair_14pm(callback):
    file_path1 = './ppg_repair.png'
    with open(file_path1, 'rb') as file1:
        keyboard = types.InlineKeyboardMarkup()

        display_orig_btn_14pm = types.InlineKeyboardButton(text='Дисплей Оригинал 🔥', callback_data='display_orig_14pm')
        display_dubl_btn_14pm = types.InlineKeyboardButton(text='Дисплей Дубликат 🔥', callback_data='display_dubl_14pm')
        steklo_btn_14pm = types.InlineKeyboardButton(text='Стекло 🔥', callback_data='steklo_14pm')
        zad_btn_14pm = types.InlineKeyboardButton(text='Задняя крышка 🔥', callback_data='zad_14pm')
        korpus_btn_14pm = types.InlineKeyboardButton(text='Корпус 🔥', callback_data='korpus_14pm')
        akb_btn_14pm = types.InlineKeyboardButton(text='Аккумулятор 🔥', callback_data='akb_14pm')
        plata_btn_14pm = types.InlineKeyboardButton(text='Плата 🔥', callback_data='plata_14pm')
        steklo_kam_btn_14pm = types.InlineKeyboardButton(text='Стекло камеры 🔥', callback_data='steklo_kam_14pm')
        kamera_zad_btn_14pm = types.InlineKeyboardButton(text='Задняя камера 🔥', callback_data='kamera_zad_14pm')
        kamera_front_btn_14pm = types.InlineKeyboardButton(text='Фронтальная камера 🔥',
                                                           callback_data='kamera_front_14pm')
        faceid_btn_14pm = types.InlineKeyboardButton(text='Face ID 🔥', callback_data='faceid_14pm')
        niz_shleif_btn_14pm = types.InlineKeyboardButton(text='Нижний шлейф 🔥', callback_data='niz_shleif_14pm')
        price_menu_back = types.InlineKeyboardButton(text='◀️ Назад', callback_data='price_menu_back')

        keyboard.row(display_orig_btn_14pm)
        keyboard.row(display_dubl_btn_14pm)
        keyboard.row(steklo_btn_14pm)
        keyboard.row(zad_btn_14pm)
        keyboard.row(korpus_btn_14pm)
        keyboard.row(akb_btn_14pm)
        keyboard.row(plata_btn_14pm)
        keyboard.row(steklo_kam_btn_14pm)
        keyboard.row(kamera_zad_btn_14pm)
        keyboard.row(kamera_front_btn_14pm)
        keyboard.row(faceid_btn_14pm)
        keyboard.row(niz_shleif_btn_14pm)
        keyboard.row(price_menu_back)

        bot.send_photo(callback.message.chat.id, file1)
        bot.send_message(callback.message.chat.id, 'Ремонтные услуги iPhone 14 Pro Max 🔥', reply_markup=keyboard)



############ ПЕРВАЯ СТРАНИЦА IPHONE УСТРОЙСТВ ##############



def first_page_iphones_repair(callback):
    file_path1 = './ppg_repair.png'
    with open(file_path1, 'rb') as file1:
        keyboard = types.InlineKeyboardMarkup()

        iphone_btn1 = types.InlineKeyboardButton(text='14 Pro Max 🔥', callback_data='repair_14pm')
        iphone_btn2 = types.InlineKeyboardButton(text='14 Pro 🔥', callback_data='repair_14pro')
        iphone_btn3 = types.InlineKeyboardButton(text='14 Plus 🔥', callback_data='repair_14plus')
        iphone_btn4 = types.InlineKeyboardButton(text='14 🔥', callback_data='repair_14')
        iphone_btn5 = types.InlineKeyboardButton(text='13PM 🔥', callback_data='repair_13pm')
        iphone_btn6 = types.InlineKeyboardButton(text='13 Pro 🔥', callback_data='repair_13pro')
        iphone_btn7 = types.InlineKeyboardButton(text='13 mini 🔥', callback_data='repair_13mini')
        iphone_btn8 = types.InlineKeyboardButton(text='13 🔥', callback_data='repair_13')
        iphone_btn9 = types.InlineKeyboardButton(text='12 Pro Max 🔥', callback_data='repair_12pm')
        iphone_btn10 = types.InlineKeyboardButton(text='12 Pro 🔥', callback_data='repair_12pro')
        iphone_btn11 = types.InlineKeyboardButton(text='12 mini 🔥', callback_data='repair_12mini')
        iphone_btn12 = types.InlineKeyboardButton(text='12 🔥', callback_data='repair_12')
        iphone_btn13 = types.InlineKeyboardButton(text='◀️ Назад', callback_data='repair_previous_page1')
        iphone_btn14 = types.InlineKeyboardButton(text='Вперед ➡️', callback_data='repair_next_page1')
        iphone_btn15 = types.InlineKeyboardButton(text='В главное меню 🔙', callback_data='repair_main_menu')

        keyboard.add(iphone_btn1, iphone_btn2)
        keyboard.add(iphone_btn3, iphone_btn4)
        keyboard.add(iphone_btn5, iphone_btn6)
        keyboard.add(iphone_btn7, iphone_btn8)
        keyboard.add(iphone_btn9, iphone_btn10)
        keyboard.add(iphone_btn11, iphone_btn12)
        keyboard.add(iphone_btn13, iphone_btn14)
        keyboard.add(iphone_btn15)

        bot.send_photo(callback.message.chat.id, file1)
        bot.send_message(callback.message.chat.id, 'Список всех устройств, обслуживающихся в PaPaGadget 🔥',
                         reply_markup=keyboard)


######## ВТОРАЯ СТРАНИЦА IPHONE УСТРОЙСТВ


def second_page_iphones_repair(callback):
    file_path1 = './ppg_repair.png'
    with open(file_path1, 'rb') as file1:
        keyboard = types.InlineKeyboardMarkup()

        iphone_btn16 = types.InlineKeyboardButton(text='11 Pro Max 🔥', callback_data='repair_11pm')
        iphone_btn17 = types.InlineKeyboardButton(text='11 Pro 🔥', callback_data='repair_11pro')
        iphone_btn18 = types.InlineKeyboardButton(text='11 🔥', callback_data='repair_11')
        iphone_btn19 = types.InlineKeyboardButton(text='XR 🔥', callback_data='repair_xr')
        iphone_btn20 = types.InlineKeyboardButton(text='XS Max 🔥', callback_data='repair_xsmax')
        iphone_btn21 = types.InlineKeyboardButton(text='XS 🔥', callback_data='repair_xs')
        iphone_btn22 = types.InlineKeyboardButton(text='X 🔥', callback_data='repair_x')
        iphone_btn23 = types.InlineKeyboardButton(text='8 Plus 🔥', callback_data='repair_8plus')
        iphone_btn24 = types.InlineKeyboardButton(text='8 🔥', callback_data='repair_8')
        iphone_btn25 = types.InlineKeyboardButton(text='7 Plus 🔥', callback_data='repair_7plus')
        iphone_btn26 = types.InlineKeyboardButton(text='7 🔥', callback_data='repair_7')
        iphone_btn27 = types.InlineKeyboardButton(text='6S Plus🔥', callback_data='repair_6splus')
        iphone_btn28 = types.InlineKeyboardButton(text='◀️ Назад', callback_data='repair_previous_page2')
        iphone_btn29 = types.InlineKeyboardButton(text='Вперед ➡️', callback_data='repair_next_page2')
        iphone_btn30 = types.InlineKeyboardButton(text='В главное меню 🔙', callback_data='repair_main_menu')

        keyboard.add(iphone_btn16, iphone_btn17)
        keyboard.add(iphone_btn18, iphone_btn19)
        keyboard.add(iphone_btn20, iphone_btn21)
        keyboard.add(iphone_btn22, iphone_btn23)
        keyboard.add(iphone_btn24, iphone_btn25)
        keyboard.add(iphone_btn26, iphone_btn27)
        keyboard.add(iphone_btn28, iphone_btn29)
        keyboard.add(iphone_btn30)

        bot.send_photo(callback.message.chat.id, file1)
        bot.send_message(callback.message.chat.id, 'Список всех устройств, обслуживающихся в PaPaGadget 🔥🔥',
                         reply_markup=keyboard)



def first_page_iphones_pricing(callback):
    file_path1 = './ppg_pricing.png'
    with open(file_path1, 'rb') as file1:
        keyboard = types.InlineKeyboardMarkup()

        iphone_btn1 = types.InlineKeyboardButton(text='14 Pro Max 🔥', callback_data='pricing_14pm')
        iphone_btn2 = types.InlineKeyboardButton(text='14 Pro 🔥', callback_data='pricing_14pro')
        iphone_btn3 = types.InlineKeyboardButton(text='14 Plus 🔥', callback_data='pricing_14plus')
        iphone_btn4 = types.InlineKeyboardButton(text='14 🔥', callback_data='pricing_14')
        iphone_btn5 = types.InlineKeyboardButton(text='13PM 🔥', callback_data='pricing_13pm')
        iphone_btn6 = types.InlineKeyboardButton(text='13 Pro 🔥', callback_data='pricing_13pro')
        iphone_btn7 = types.InlineKeyboardButton(text='13 mini 🔥', callback_data='pricing_13mini')
        iphone_btn8 = types.InlineKeyboardButton(text='13 🔥', callback_data='pricing_13')
        iphone_btn9 = types.InlineKeyboardButton(text='12 Pro Max 🔥', callback_data='pricing_12pm')
        iphone_btn10 = types.InlineKeyboardButton(text='12 Pro 🔥', callback_data='pricing_12pro')
        iphone_btn11 = types.InlineKeyboardButton(text='12 mini 🔥', callback_data='pricing_12mini')
        iphone_btn12 = types.InlineKeyboardButton(text='12 🔥', callback_data='pricing_12')
        iphone_btn13 = types.InlineKeyboardButton(text='◀️ Назад', callback_data='pricing_previous_page1')
        iphone_btn14 = types.InlineKeyboardButton(text='Вперед ➡️', callback_data='pricing_next_page1')

        keyboard.add(iphone_btn1, iphone_btn2)
        keyboard.add(iphone_btn3, iphone_btn4)
        keyboard.add(iphone_btn5, iphone_btn6)
        keyboard.add(iphone_btn7, iphone_btn8)
        keyboard.add(iphone_btn9, iphone_btn10)
        keyboard.add(iphone_btn11, iphone_btn12)
        keyboard.add(iphone_btn13, iphone_btn14)

        bot.send_photo(callback.message.chat.id, file1)
        bot.send_message(callback.message.chat.id, 'Список всех устройств, которые Вы можете продать в PaPaGadget 🔥',
                         reply_markup=keyboard)


#### ОЦЕНКА 14 про макс начало



def pricing_14pm_0step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path = './ppg_tradein.png'
    with open(file_path, 'rb') as file:
        keyboard = types.InlineKeyboardMarkup()

        korpus_14pm_btn1 = types.InlineKeyboardButton(text='В идеале 🔥', callback_data='korpus_14pm_ideal')
        korpus_14pm_btn2 = types.InlineKeyboardButton(text='Немного потертый ✅', callback_data='korpus_14pm_medium')
        korpus_14pm_btn3 = types.InlineKeyboardButton(text='В царапинах / поношенный 💀',
                                                      callback_data='korpus_14pm_bad')

        keyboard.add(korpus_14pm_btn1)
        keyboard.add(korpus_14pm_btn2)
        keyboard.add(korpus_14pm_btn3)

        bot.send_photo(callback.message.chat.id, file)
        bot.send_message(callback.message.chat.id, 'В каком состоянии Ваш корпус? 🔥', reply_markup=keyboard)



########### ОЦЕНКА КОРПУСА 14 ПМ ПЕРЕХОД НА ОЦЕНКУ ДИСПЛЕЯ #######
def pricing_14pm_1step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path = './ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard = types.InlineKeyboardMarkup()
        screen_14pm_btn1 = types.InlineKeyboardButton(text='В идеале 🔥', callback_data='display_14pm_ideal')
        screen_14pm_btn2 = types.InlineKeyboardButton(text='Царапины на экране ✅', callback_data='display_14pm_scratch')
        screen_14pm_btn3 = types.InlineKeyboardButton(text='Разбито стекло дисплея 💀',
                                                      callback_data='display_14pm_badglass')
        screen_14pm_btn4 = types.InlineKeyboardButton(text='Полосы / пятна на экране 💀',
                                                      callback_data='display_14pm_badlines')
        screen_14pm_btn5 = types.InlineKeyboardButton(text='Экран не работает 💀', callback_data='display_14pm_fucked')

        keyboard.add(screen_14pm_btn1)
        keyboard.add(screen_14pm_btn2)
        keyboard.add(screen_14pm_btn3)
        keyboard.add(screen_14pm_btn4)
        keyboard.add(screen_14pm_btn5)

        bot.send_photo(callback.message.chat.id, file)
        bot.send_message(callback.message.chat.id, 'В каком состоянии Ваш экран? 🔥', reply_markup=keyboard)



####### ОЦЕНКА ДИСПЛЕЯ 14 ПМ ПЕРЕХОД НА ОЦЕНКУ КАМЕР #######

def pricing_14pm_2step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path='./ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard=types.InlineKeyboardMarkup()

        camera_14pm_btn1 = types.InlineKeyboardButton(text='В идеале 🔥', callback_data='camera_14pm_ideal')
        camera_14pm_btn2 = types.InlineKeyboardButton(text='Разбито стекло камеры ✅', callback_data='camera_14pm_badglass')
        camera_14pm_btn3 = types.InlineKeyboardButton(text='Не работает камера 💀', callback_data='camera_14pm_fucked')

        keyboard.add(camera_14pm_btn1)
        keyboard.add(camera_14pm_btn2)
        keyboard.add(camera_14pm_btn3)

        bot.send_photo(callback.message.chat.id,file)
        bot.send_message(callback.message.chat.id, 'В каком состоянии Ваша камера? 🔥', reply_markup=keyboard)



######## Оценка других проблем 14 ПМ



def pricing_14pm_3step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path='./ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard=types.InlineKeyboardMarkup()

        other_14pm_btn1 = types.InlineKeyboardButton(text='Других проблем нет 🔥', callback_data='other_14pm_ideal')
        other_14pm_btn2 = types.InlineKeyboardButton(text='Не работает Face ID ✅', callback_data='other_14pm_faceid')
        other_14pm_btn3 = types.InlineKeyboardButton(text='Проблемы со звуком ✅', callback_data='other_14pm_sound')
        other_14pm_btn4 = types.InlineKeyboardButton(text='Нерабочие кнопки ✅', callback_data='other_14pm_buttons')

        keyboard.add(other_14pm_btn1)
        keyboard.add(other_14pm_btn2)
        keyboard.add(other_14pm_btn3)
        keyboard.add(other_14pm_btn4)

        bot.send_photo(callback.message.chat.id,file)
        bot.send_message(callback.message.chat.id, 'Есть ли другие неисправности в телефоне? 🔥', reply_markup=keyboard)



###### ФИНАЛ ОЦЕНКИ 14ПМ

def pricing_14pm_4step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path = './ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard = types.InlineKeyboardMarkup()

        final_14pm_btn1 = types.InlineKeyboardButton(text='Ваш iPhone 14 Pro Max оценён :', callback_data='final_14pm_description')
        final_14pm_btn2 = types.InlineKeyboardButton(text=f'{price14pm} ✅', callback_data='final_14pm_price')
        final_14pm_btn3 = types.InlineKeyboardButton(text='В главное меню 🔙', callback_data='back_main_menu_14pm_pricing')

        keyboard.add(final_14pm_btn1)
        keyboard.add(final_14pm_btn2)
        keyboard.add(final_14pm_btn3)

        bot.send_photo(callback.message.chat.id, file, reply_markup=keyboard)




        ####### Оценка 14 про начало

def pricing_14pro_0step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path = './ppg_tradein.png'
    with open(file_path, 'rb') as file:
        keyboard = types.InlineKeyboardMarkup()

        korpus_14pm_btn1 = types.InlineKeyboardButton(text='В идеале 🔥', callback_data='korpus_14pm_ideal')
        korpus_14pm_btn2 = types.InlineKeyboardButton(text='Немного потертый ✅', callback_data='korpus_14pm_medium')
        korpus_14pm_btn3 = types.InlineKeyboardButton(text='В царапинах / поношенный 💀',
                                                      callback_data='korpus_14pm_bad')

        keyboard.add(korpus_14pm_btn1)
        keyboard.add(korpus_14pm_btn2)
        keyboard.add(korpus_14pm_btn3)

        bot.send_photo(callback.message.chat.id, file)
        bot.send_message(callback.message.chat.id, 'В каком состоянии Ваш корпус? 🔥', reply_markup=keyboard)



#####ОЦЕНКА 14 про дисплей

def pricing_14pro_1step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path = './ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard = types.InlineKeyboardMarkup()
        screen_14pro_btn1 = types.InlineKeyboardButton(text='В идеале 🔥', callback_data='display_14pro_ideal')
        screen_14pro_btn2 = types.InlineKeyboardButton(text='Царапины на экране ✅', callback_data='display_14pro_scratch')
        screen_14pro_btn3 = types.InlineKeyboardButton(text='Разбито стекло дисплея 💀',
                                                      callback_data='display_14pro_badglass')
        screen_14pro_btn4 = types.InlineKeyboardButton(text='Полосы / пятна на экране 💀',
                                                      callback_data='display_14pro_badlines')
        screen_14pro_btn5 = types.InlineKeyboardButton(text='Экран не работает 💀', callback_data='display_14pro_fucked')

        keyboard.add(screen_14pro_btn1)
        keyboard.add(screen_14pro_btn2)
        keyboard.add(screen_14pro_btn3)
        keyboard.add(screen_14pro_btn4)
        keyboard.add(screen_14pro_btn5)

        bot.send_photo(callback.message.chat.id, file)
        bot.send_message(callback.message.chat.id, 'В каком состоянии Ваш экран? 🔥', reply_markup=keyboard)



#### ОЦЕНКА 14 ПРО камеры

def pricing_14pro_2step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path='./ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard=types.InlineKeyboardMarkup()

        camera_14pro_btn1 = types.InlineKeyboardButton(text='В идеале 🔥', callback_data='camera_14pro_ideal')
        camera_14pro_btn2 = types.InlineKeyboardButton(text='Разбито стекло камеры ✅', callback_data='camera_14pro_badglass')
        camera_14pro_btn3 = types.InlineKeyboardButton(text='Не работает камера 💀', callback_data='camera_14pro_fucked')

        keyboard.add(camera_14pro_btn1)
        keyboard.add(camera_14pro_btn2)
        keyboard.add(camera_14pro_btn3)

        bot.send_photo(callback.message.chat.id,file)
        bot.send_message(callback.message.chat.id, 'В каком состоянии Ваша камера? 🔥', reply_markup=keyboard)


###### ОЦЕНКА 14 ПРО ДРУГИЕ ПРОБЛЕМЫ

def pricing_14pro_3step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path='./ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard=types.InlineKeyboardMarkup()

        other_14pro_btn1 = types.InlineKeyboardButton(text='Других проблем нет 🔥', callback_data='other_14pro_ideal')
        other_14pro_btn2 = types.InlineKeyboardButton(text='Не работает Face ID ✅', callback_data='other_14pro_faceid')
        other_14pro_btn3 = types.InlineKeyboardButton(text='Проблемы со звуком ✅', callback_data='other_14pro_sound')
        other_14pro_btn4 = types.InlineKeyboardButton(text='Нерабочие кнопки ✅', callback_data='other_14pro_buttons')

        keyboard.add(other_14pro_btn1)
        keyboard.add(other_14pro_btn2)
        keyboard.add(other_14pro_btn3)
        keyboard.add(other_14pro_btn4)

        bot.send_photo(callback.message.chat.id,file)
        bot.send_message(callback.message.chat.id, 'Есть ли другие неисправности в телефоне? 🔥', reply_markup=keyboard)


#### ОЦЕНКА 14 ПРО ФИНАЛ

def pricing_14pro_4step(callback):
    bot.delete_message(callback.message.chat.id, callback.message.message_id)
    bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
    file_path = './ppg_pricing.png'
    with open(file_path, 'rb') as file:
        keyboard = types.InlineKeyboardMarkup()

        final_14pro_btn1 = types.InlineKeyboardButton(text='Ваш iPhone 14 Pro оценён :', callback_data='final_14pro_description')
        final_14pro_btn2 = types.InlineKeyboardButton(text=f'{price14pm} ✅', callback_data='final_14pro_price')
        final_14pro_btn3 = types.InlineKeyboardButton(text='В главное меню 🔙', callback_data='back_main_menu_14pro_pricing')

        keyboard.add(final_14pro_btn1)
        keyboard.add(final_14pro_btn2)
        keyboard.add(final_14pro_btn3)

        bot.send_photo(callback.message.chat.id, file, reply_markup=keyboard)

        ####### REPLY КЛАВИАТУРА


@bot.message_handler(commands=['start'])
def start(message):

    #Основные кнопки

    markup = types.ReplyKeyboardMarkup()
    btn1 = types.KeyboardButton('Trade In ♻️')
    btn2 = types.KeyboardButton('Ремонт 🛠')
    btn3 = types.KeyboardButton('Оценка устройства 💼')
    btn4 = types.KeyboardButton('Аксессуары ☎️')
    btn5 = types.KeyboardButton('Original 🍏')
    btn6 = types.KeyboardButton('Связаться с нами 💬')
    markup.row(btn1, btn2)
    markup.row(btn4,btn5)
    markup.row(btn3, btn6)




    #Приветствие

    bot.send_message(message.chat.id,'Доброго времени суток уважаемый Клиент! Этот бот - Ваш личный помощник в сервисе PaPaGadget 🔥', reply_markup=markup)


###### СВЯЗАТЬСЯ С НАМИ #####
@bot.message_handler(func=lambda message:message.text == 'Связаться с нами 💬')
def support(message):
    if message.text=='Связаться с нами 💬':
        file_path = './ppg_contactus.png'
        with open(file_path, 'rb') as file:
            markup = types.InlineKeyboardMarkup()
            support_btn = types.InlineKeyboardButton('Напишите нам 💬', url='https://t.me/molyrekt')
            markup.add(support_btn)
            bot.send_photo(message.chat.id,file, reply_markup=markup)




#Trade In Menu

@bot.message_handler(func=lambda message: message.text == 'Trade In ♻️')
def tradein(message):
    if message.text == 'Trade In ♻️':
        file_path = './ppg_tradein.png'
        with open(file_path, 'rb') as image_file:
            markup = types.InlineKeyboardMarkup()

            with open('tradein.csv', 'r') as csv_file:
                csv_reader = csv.reader(csv_file)
                for row in csv_reader:
                    _, button_text, callback_data = row
                    button = types.InlineKeyboardButton(button_text, callback_data=callback_data)
                    markup.add(button)

            bot.send_photo(message.chat.id, image_file, reply_markup=markup)

price14pm = 500000
price14pro = 480000




############Tradein MENU


# Загрузка книги Excel
wb = openpyxl.load_workbook('path_to_your_excel_file.xlsx', data_only=True)
ws = wb.active  # Активный лист

# Создание структуры данных для хранения информации о каждом устройстве
devices_info = []

# Загрузчик изображений из Excel
image_loader = SheetImageLoader(ws)

for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
    device_info = {
        "name": row[1],
        "callback_data": row[2],
        "buttons": [
            {"text": row[3], "callback_data": "x"},
            {"text": row[4], "callback_data": "x"},
            {"text": row[5], "callback_data": "x"},
            {"text": row[6], "callback_data": "x"}
        ],
        "back_button": {"text": "Назад ◀️", "callback_data": "x"},
        "images": [image_loader.get(f'I{row[0]}'), image_loader.get(f'J{row[0]}')]  # Предполагается, что изображения находятся в столбцах I и J
    }
    devices_info.append(device_info)

@bot.callback_query_handler(func=lambda callback: callback.data)
def callback_tradein(callback):
    for device in devices_info:
        if callback.data == device["callback_data"]:
            bot.delete_message(callback.message.chat.id, callback.message.message_id)
            keyboard = types.InlineKeyboardMarkup()

            # Добавление кнопок для устройства
            for btn_info in device["buttons"]:
                btn = types.InlineKeyboardButton(text=btn_info["text"], callback_data=btn_info["callback_data"])
                keyboard.add(btn)

            # Добавление кнопки 'Назад'
            back_btn_info = device["back_button"]
            back_btn = types.InlineKeyboardButton(text=back_btn_info["text"], callback_data=back_btn_info["callback_data"])
            keyboard.add(back_btn)

            # Подготовка и отправка изображений
            media = [types.InputMediaPhoto(media=img) for img in device["images"] if img is not None]

            bot.send_media_group(callback.message.chat.id, media=media)
            bot.send_message(callback.message.chat.id, device["name"], reply_markup=keyboard)


    global price14pm, price14pro




############Возвращение в трейд ин меню


    if callback.data=='back_tradein':
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
        bot.delete_message(callback.message.chat.id, callback.message.message_id - 2)

        main_menu_tradein(callback)


        # СПИСОК IPHONE УСТРОЙСТВ


    if callback.data == 'iphone':
            bot.delete_message(callback.message.chat.id, callback.message.message_id)
            bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)

            first_page_iphones_repair(callback)



    if callback.data=='repair_main_menu':
        file_path1='./ppg_repair.png'
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
        with open(file_path1, 'rb') as file:
            markup = types.InlineKeyboardMarkup()
            repair_btn2 = types.InlineKeyboardButton('iPhone 🔥', callback_data='iphone')
            repair_btn3 = types.InlineKeyboardButton('Samsung 🔥', callback_data='samsung')
            repair_btn4 = types.InlineKeyboardButton('Xiaomi 🔥', callback_data='xiaomi')
            repair_btn5 = types.InlineKeyboardButton('Poco 🔥', callback_data='poco')
            markup.row(repair_btn2, repair_btn3)
            markup.row(repair_btn4, repair_btn5)
            bot.send_photo(callback.message.chat.id, file)
            bot.send_message(callback.message.chat.id, '📱 Выберите Ваше устройство 📱',
                             reply_markup=markup)




############## Переход с первой страницы на вторую



    if callback.data=='repair_next_page1':
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)

        second_page_iphones_repair(callback)






############## Возвращение со второй страницы на первую



    if callback.data=='repair_previous_page2':
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
        first_page_iphones_repair(callback)


            ####### Переход со второй страницы на третью

    if callback.data=='repair_next_page2':
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
        file_path1 = './ppg_tradein.png'
        with open(file_path1, 'rb') as file1:
            keyboard = types.InlineKeyboardMarkup()

            iphone_btn31 = types.InlineKeyboardButton(text='6S 🔥', callback_data='repair_6s')
            iphone_btn32 = types.InlineKeyboardButton(text='6 Plus🔥', callback_data='repair_6plus')
            iphone_btn33 = types.InlineKeyboardButton(text='6 🔥', callback_data='repair_6')
            iphone_btn34 = types.InlineKeyboardButton(text='SE 2020 🔥', callback_data='repair_se2020')
            iphone_btn35 = types.InlineKeyboardButton(text='◀️ Назад', callback_data='repair_previous_page3')
            iphone_btn36 = types.InlineKeyboardButton(text='Вперед ➡️', callback_data='repair_next_page3')
            iphone_btn37 = types.InlineKeyboardButton(text='В главное меню 🔙', callback_data='repair_main_menu')


            keyboard.add(iphone_btn31, iphone_btn32)
            keyboard.add(iphone_btn33, iphone_btn34)
            keyboard.add(iphone_btn35, iphone_btn36)
            keyboard.add(iphone_btn37)


            bot.send_photo(callback.message.chat.id, file1)
            bot.send_message(callback.message.chat.id, 'Список всех устройств, обслуживающихся в PaPaGadget 🔥🔥', reply_markup=keyboard)


###### Возвращение с третьей страницы на вторую


    if callback.data=='repair_previous_page3':
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)

        second_page_iphones_repair(callback)



            ############  Прайс лист 14 pro max


    if callback.data=='repair_14pm':
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)

        main_menu_repair_14pm(callback)


######## ИНФО О ОРИГИНАЛ ДИСПЛЕЕ 14 ПРО МАКС



    if callback.data=='display_orig_14pm':
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)
        file_path1='./ppg_repair.png'
        with open(file_path1,'rb') as file1:
            keyboard = types.InlineKeyboardMarkup()

            display_orig_14pm_price = types.InlineKeyboardButton(text = '💰 Цена : 240 000 тг', callback_data='display_orig_14pm_price')
            display_orig_14pm_guarantee = types.InlineKeyboardButton(text = '✅ Гарантия : 12 месяцев', callback_data='display_orig_14pm_guarantee')
            display_orig_14pm_back = types.InlineKeyboardButton(text = '◀️ Назад', callback_data='display_orig_14pm_back')

            keyboard.row(display_orig_14pm_price)
            keyboard.row(display_orig_14pm_guarantee)
            keyboard.row(display_orig_14pm_back)

            bot.send_photo(callback.message.chat.id, file1, reply_markup=keyboard)


######### Возвращение с инфы о ремонте на список ремонтных услуг по айфону

    if callback.data=='display_orig_14pm_back':
        bot.delete_message(callback.message.chat.id, callback.message.message_id)

        main_menu_repair_14pm(callback)


    ########## Возвращение в меню ремонта


    if callback.data=='price_menu_back':
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        bot.delete_message(callback.message.chat.id, callback.message.message_id - 1)

        first_page_iphones_repair(callback)









    ######## Оценка 14 про макс


    if callback.data=='pricing_14pm':
        pricing_14pm_0step(callback)


##### ОЦЕНКА КОРПУСА 14 ПМ ИДЕАЛ


    if callback.data == 'korpus_14pm_ideal':
        price14pm-=0
        pricing_14pm_1step(callback)


###### ОЦЕНКА КОРПУСА 14 ПМ МЕДИУМ


    elif callback.data == 'korpus_14pm_medium':
        price14pm-=40000
        pricing_14pm_1step(callback)


###### ОЦЕНКА КОРПУСА 14 ПМ УГАШЕН


    elif callback.data == 'korpus_14pm_bad':
        price14pm-=65000
        pricing_14pm_1step(callback)




###### ОЦЕНКА ДИСПЛЕЯ 14 ПМ ИДЕАЛ

    elif callback.data == 'display_14pm_ideal':
        price14pm-=0
        pricing_14pm_2step(callback)



##### ОЦЕНКА ДИСПЛЕЯ 14 ПМ МЕДИУМ

    elif callback.data == 'display_14pm_scratch':
        price14pm-=10000
        pricing_14pm_2step(callback)


###### ОЦЕНКА ДИСПЛЕЯ 14 ПМ РАЗБИТО СТЕКЛО

    elif callback.data == 'display_14pm_badglass':
        price14pm-=30000
        pricing_14pm_2step(callback)


####### ОЦЕНКА ДИСПЛЕЯ 14 ПМ ПЯТНА ПОЛОСЫ

    elif callback.data == 'display_14pm_badlines':
        price14pm-=200000
        pricing_14pm_2step(callback)


###### ОЦЕНКА ДИСПЛЕЯ 14 ПМ НЕ РАБОТАЕТ ЭКРАН

    elif callback.data == 'display_14pm_fucked':
        price14pm-=230000
        pricing_14pm_2step(callback)

###### ОЦЕНКА КАМЕРЫ 14 ПМ ИДЕАЛ

    elif callback.data == 'camera_14pm_ideal':
        price14pm-=0
        pricing_14pm_3step(callback)

###### ОЦЕНКА КАМЕРЫ 14 ПМ РАЗБИТО СТЕКЛО КАМЕРЫ

    elif callback.data == 'camera_14pm_badglass':
        price14pm-=10000
        pricing_14pm_3step(callback)

####### ОЦЕНКА КАМЕРЫ 14ПМ НЕ РАБОТАЕТ КАМЕРА

    elif callback.data == 'camera_14pm_fucked':
        price14pm-=25000
        pricing_14pm_3step(callback)


####### Оценка других проблем 14 пм Нет проблем

    elif callback.data == 'other_14pm_ideal':
        price14pm-=0
        pricing_14pm_4step(callback)

###### Оценка других проблем 14 пм face id

    elif callback.data == 'other_14pm_faceid':
        price14pm-=30000
        pricing_14pm_4step(callback)

        ######Оценка других проблем 14пм проблемы со звуком

    elif callback.data == 'other_14pm_sound':
        price14pm-=15000
        pricing_14pm_4step(callback)


        ###### Оценка других проблем 14пм проблемы с кнопками

    elif callback.data == 'other_14pm_buttons':
        price14pm-=15000
        pricing_14pm_4step(callback)

####### Возврат в главное меню из 14пм

    elif callback.data == 'back_main_menu_14pm_pricing':
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        first_page_iphones_pricing(callback)
        price14pm=500000





    ######## Оценка 14 про макс


    if callback.data=='pricing_14pro':
        pricing_14pro_0step(callback)


##### ОЦЕНКА КОРПУСА 14 Про ИДЕАЛ


    if callback.data == 'korpus_14pro_ideal':
        price14pro-=0
        pricing_14pro_1step(callback)


###### ОЦЕНКА КОРПУСА 14 Про МЕДИУМ


    elif callback.data == 'korpus_14pro_medium':
        price14pro-=42000
        pricing_14pro_1step(callback)


###### ОЦЕНКА КОРПУСА 14 Про УГАШЕН


    elif callback.data == 'korpus_14pro_bad':
        price14pro-=60000
        pricing_14pro_1step(callback)




###### ОЦЕНКА ДИСПЛЕЯ 14 Про ИДЕАЛ

    elif callback.data == 'display_14pro_ideal':
        price14pro-=0
        pricing_14pro_2step(callback)



##### ОЦЕНКА ДИСПЛЕЯ 14 Про МЕДИУМ

    elif callback.data == 'display_14pro_scratch':
        price14pro-=12000
        pricing_14pro_2step(callback)


###### ОЦЕНКА ДИСПЛЕЯ 14 Про РАЗБИТО СТЕКЛО

    elif callback.data == 'display_14pro_badglass':
        price14pro-=35000
        pricing_14pro_2step(callback)


####### ОЦЕНКА ДИСПЛЕЯ 14 Про ПЯТНА ПОЛОСЫ

    elif callback.data == 'display_14pro_badlines':
        price14pro-=210000
        pricing_14pro_2step(callback)


###### ОЦЕНКА ДИСПЛЕЯ 14 Про НЕ РАБОТАЕТ ЭКРАН

    elif callback.data == 'display_14pro_fucked':
        price14pro-=220000
        pricing_14pro_2step(callback)

###### ОЦЕНКА КАМЕРЫ 14 Про ИДЕАЛ

    elif callback.data == 'camera_14pro_ideal':
        price14pro-=0
        pricing_14pro_3step(callback)

###### ОЦЕНКА КАМЕРЫ 14 Про РАЗБИТО СТЕКЛО КАМЕРЫ

    elif callback.data == 'camera_14pro_badglass':
        price14pro-=12000
        pricing_14pro_3step(callback)

####### ОЦЕНКА КАМЕРЫ 14Про НЕ РАБОТАЕТ КАМЕРА

    elif callback.data == 'camera_14pro_fucked':
        price14pro-=20000
        pricing_14pro_3step(callback)


####### Оценка других проблем 14 про Нет проблем

    elif callback.data == 'other_14pro_ideal':
        price14pro-=0
        pricing_14pro_4step(callback)

###### Оценка других проблем 14 про face id

    elif callback.data == 'other_14pro_faceid':
        price14pro-=25000
        pricing_14pro_4step(callback)

        ######Оценка других проблем 14про проблемы со звуком

    elif callback.data == 'other_14pro_sound':
        price14pro-=10000
        pricing_14pro_4step(callback)


        ###### Оценка других проблем 14про проблемы с кнопками

    elif callback.data == 'other_14pro_buttons':
        price14pro-=10000
        pricing_14pro_4step(callback)

####### Возврат в главное меню из 14про

    elif callback.data == 'back_main_menu_14pro_pricing':
        bot.delete_message(callback.message.chat.id, callback.message.message_id)
        first_page_iphones_pricing(callback)
        price14pro=480000


########### Ремонт - основной раздел

@bot.message_handler(func=lambda message: message.text == 'Ремонт 🛠')
def repair(message):
    if message.text == "Ремонт 🛠":
        file_path = './ppg_repair.png'
        with open(file_path, 'rb') as file:
            markup = types.InlineKeyboardMarkup()
            repair_btn1 = types.InlineKeyboardButton('Выберите Ваше устройство 📱',callback_data='choose')
            repair_btn2 = types.InlineKeyboardButton('iPhone 🔥', callback_data='iphone')
            repair_btn3 = types.InlineKeyboardButton('Samsung 🔥', callback_data='samsung')
            repair_btn4 = types.InlineKeyboardButton('Xiaomi 🔥', callback_data='xiaomi')
            repair_btn5 = types.InlineKeyboardButton('Poco 🔥', callback_data='poco')
            markup.row(repair_btn1)
            markup.row(repair_btn2,repair_btn3)
            markup.row(repair_btn4,repair_btn5)
            bot.send_photo(message.chat.id,file,reply_markup=markup)
        file.close()


@bot.message_handler(func=lambda message: message.text == 'Оценка устройства 💼')
def iphone_pricing(message):
    if message.text == "Оценка устройства 💼":
        file_path1 = './ppg_pricing.png'
        with open(file_path1, 'rb') as file1:
            keyboard = types.InlineKeyboardMarkup()

            iphone_btn1 = types.InlineKeyboardButton(text='14 Pro Max 🔥', callback_data='pricing_14pm')
            iphone_btn2 = types.InlineKeyboardButton(text='14 Pro 🔥', callback_data='pricing_14pro')
            iphone_btn3 = types.InlineKeyboardButton(text='14 Plus 🔥', callback_data='pricing_14plus')
            iphone_btn4 = types.InlineKeyboardButton(text='14 🔥', callback_data='pricing_14')
            iphone_btn5 = types.InlineKeyboardButton(text='13PM 🔥', callback_data='pricing_13pm')
            iphone_btn6 = types.InlineKeyboardButton(text='13 Pro 🔥', callback_data='pricing_13pro')
            iphone_btn7 = types.InlineKeyboardButton(text='13 mini 🔥', callback_data='pricing_13mini')
            iphone_btn8 = types.InlineKeyboardButton(text='13 🔥', callback_data='pricing_13')
            iphone_btn9 = types.InlineKeyboardButton(text='12 Pro Max 🔥', callback_data='pricing_12pm')
            iphone_btn10 = types.InlineKeyboardButton(text='12 Pro 🔥', callback_data='pricing_12pro')
            iphone_btn11 = types.InlineKeyboardButton(text='12 mini 🔥', callback_data='pricing_12mini')
            iphone_btn12 = types.InlineKeyboardButton(text='12 🔥', callback_data='pricing_12')
            iphone_btn13 = types.InlineKeyboardButton(text='◀️ Назад', callback_data='pricing_previous_page1')
            iphone_btn14 = types.InlineKeyboardButton(text='Вперед ➡️', callback_data='pricing_next_page1')

            keyboard.add(iphone_btn1, iphone_btn2)
            keyboard.add(iphone_btn3, iphone_btn4)
            keyboard.add(iphone_btn5, iphone_btn6)
            keyboard.add(iphone_btn7, iphone_btn8)
            keyboard.add(iphone_btn9, iphone_btn10)
            keyboard.add(iphone_btn11, iphone_btn12)
            keyboard.add(iphone_btn13, iphone_btn14)

            bot.send_photo(message.chat.id, file1)
            bot.send_message(message.chat.id, 'Список всех устройств, которые Вы можете продать в PaPaGadget 🔥',
                             reply_markup=keyboard)


bot.polling(none_stop=True)